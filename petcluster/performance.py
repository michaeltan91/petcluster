'''Contains the performance calculation and visualization'''
from collections import Counter
import re
import openpyxl
import plotly.graph_objects as go
import plotly.express as px

from plotly.subplots import make_subplots
import numpy as np

class Performance(object):
    """Main performance object for calculating cluster performance and sankey diagrams"""

    def __init__(self, multiplex, nodes, process_nodes, component_dict):

        self.multiplex = multiplex
        self.nodes = nodes
        self.process_nodes = process_nodes
        self.component_dict = component_dict


    def carbon_sankey(self, process_list="", ignore_list="", cutoff=0.1, fig_width=1500,\
                    fig_height=750, fig_pad=150, fig_thickness = 10, text_font = 30, \
                    title=True, process_abrev = False):
        '''Depicts all the carbon flows in the cluster in a sankey diagram'''
        label = []

        data = {}
        data['source'] = []
        data['target'] = []
        data['value'] = []
        data['color'] = []
        data['label'] = []
        counter = 0

        if process_abrev is True:
            process_name = 'name_abbrev'
        else:
            process_name = 'name'


        for link in self.multiplex.get_edges():
            if not process_list:
                if link[0][1] == 'Material' and link[1][1] == 'Material':
                    stream = self.multiplex[link[0]][link[1]][link[2]]
                    carbon_content = stream['carbon_content']*stream['mass_flow_rate']

                    if carbon_content >= cutoff and link[0][0] not in ignore_list:
                        try:
                            name_source = self.nodes[link[0][0]][process_name]
                        except KeyError:
                            try:
                                name_source = self.nodes[link[0][0]]['name']
                            except KeyError:
                                name_source = link[0][0]

                        try:
                            name_target = self.nodes[link[1][0]][process_name]
                        except KeyError:
                            try:
                                name_target = self.nodes[link[1][0]]['name']
                            except KeyError:
                                name_target = link[1][0]

                        if name_target == 'Stack':
                            name_target = 'Environment'

                        if name_source in label and name_target in label:
                            indices1 = [i for i, x in enumerate(data['source']) \
                                        if x == label.index(name_source)]
                            indices2 = [i for i, x in enumerate(data['target']) \
                                        if x == label.index(name_target)]
                            intersect=set(indices1).intersection(indices2)
                            if len(intersect) > 0:
                                data['value'][intersect.pop()] += carbon_content
                            else:
                                if name_source not in label:
                                    label.append(name_source)
                                    data['source'].append(counter)
                                    counter += 1
                                else:
                                    label.index(name_source)
                                    data['source'].append(label.index(name_source))


                                if name_target not in label:
                                    label.append(name_target)
                                    data['target'].append(counter)
                                    counter += 1
                                else:
                                    data['target'].append(label.index(name_target))

                                data['value'].append(carbon_content)

                        else:
                            if name_source not in label:
                                label.append(name_source)
                                data['source'].append(counter)
                                counter += 1
                            else:
                                label.index(name_source)
                                data['source'].append(label.index(name_source))


                            if name_target not in label:
                                label.append(name_target)
                                data['target'].append(counter)
                                counter += 1
                            else:
                                data['target'].append(label.index(name_target))

                            data['value'].append(carbon_content)
                    color = list(np.random.choice(range(256), size=3))
                    temp = f"rgba({color[0]},{color[1]},{color[2]},0.5)"
                    data['color'].append(temp)

            else:
                if link[0][1] == 'Material' and link[1][1] == 'Material':
                    if link[0][0] in process_list or link[1][0] in process_list:

                        stream = self.multiplex[link[0]][link[1]][link[2]]
                        carbon_content = stream['carbon_content']*stream['mass_flow_rate']

                        if carbon_content >= cutoff and link[0][0] not in ignore_list:

                            try:
                                name_source = self.nodes[link[0][0]][process_name]
                            except KeyError:
                                try:
                                    name_source = self.nodes[link[0][0]]['name']
                                except KeyError:
                                    name_source = link[0][0]

                            try:
                                name_target = self.nodes[link[1][0]][process_name]
                            except KeyError:
                                try:
                                    name_target = self.nodes[link[1][0]]['name']
                                except KeyError:
                                    name_target = link[1][0]

                            if name_target == 'Stack':
                                name_target = 'Environment'

                            if name_source in label and name_target in label:
                                indices1 = [i for i, x in enumerate(data['source']) \
                                            if x == label.index(name_source)]
                                indices2 = [i for i, x in enumerate(data['target']) \
                                            if x == label.index(name_target)]
                                intersect=set(indices1).intersection(indices2)
                                if len(intersect) > 0:
                                    data['value'][intersect.pop()] += carbon_content

                                else:
                                    if name_source not in label:
                                        label.append(name_source)
                                        data['source'].append(counter)
                                        counter += 1
                                    else:
                                        label.index(name_source)
                                        data['source'].append(label.index(name_source))

                                    if name_target not in label:
                                        label.append(name_target)
                                        data['target'].append(counter)
                                        counter += 1
                                    else:
                                        data['target'].append(label.index(name_target))

                                    data['value'].append(carbon_content)

                            else:

                                if name_source not in label:
                                    label.append(name_source)
                                    data['source'].append(counter)
                                    counter += 1
                                else:
                                    label.index(name_source)
                                    data['source'].append(label.index(name_source))


                                if name_target not in label:
                                    label.append(name_target)
                                    data['target'].append(counter)
                                    counter += 1
                                else:
                                    data['target'].append(label.index(name_target))

                                data['value'].append(carbon_content)
                        color = list(np.random.choice(range(256), size=3))
                        temp = f"rgba({color[0]},{color[1]},{color[2]},0.5)"
                        data['color'].append(temp)

        fig = go.Figure(data=[go.Sankey(
            valueformat = ".2f",
            valuesuffix = "ktonne/oper-year",
            node = dict(
            pad = fig_pad,
            thickness = fig_thickness,
            line = dict(color = "black", width = 1.5),
            label = label,
            color = "blue"
            ),
            link = data
        )])

        if title is True:
            fig.update_layout(title_text="Carbon flows", font_size=text_font,
            autosize = False,
            width = fig_width,
            height = fig_height)
        else:
            fig.update_layout(font_size=text_font,
            autosize = False,
            width = fig_width,
            height = fig_height)

        config = {
        'toImageButtonOptions': { 'height': None, 'width': None}
        }
        fig.show(config=config)
        fig.write_image("carbon_sankey.pdf")


    def carbon_sankey_v2(self, process_list="", ignore_list="", cutoff=0.1, fig_width=1500,\
                    fig_height=750, fig_pad=150, fig_thickness = 10, text_font = 30, bold=False,  \
                    title=True, process_abrev = False, process_uid=False, print_excel=False):
        '''Depicts all the carbon flows in the cluster in a sankey diagram'''
        label = []

        data = {}
        data['source'] = []
        data['target'] = []
        data['value'] = []
        data['color'] = []
        data['label'] = []

        node_color = []
        counter = 0

        if process_abrev is True:
            if process_uid is True:
                process_name = "source"
            else:
                process_name = 'name_abbrev'
        else:
            process_name = 'name'


        for link in self.multiplex.get_edges():
            if not process_list:
                if link[0][1] == 'Material' and link[1][1] == 'Material':
                    stream = self.multiplex[link[0]][link[1]][link[2]]
                    carbon_content = stream['carbon_content']*stream['mass_flow_rate']

                    # if carbon_content >= cutoff and link[0][0] not in ignore_list:
                    if link[0][0] not in ignore_list:
                        try:
                            name_source = self.nodes[link[0][0]][process_name]
                        except KeyError:
                            try:
                                name_source = self.nodes[link[0][0]]['name']
                            except KeyError:
                                name_source = link[0][0]

                        try:
                            name_target = self.nodes[link[1][0]][process_name]
                        except KeyError:
                            try:
                                name_target = self.nodes[link[1][0]]['name']
                            except KeyError:
                                name_target = link[1][0]

                        if name_target == 'Stack':
                            name_target = 'Environment'

                        if name_source in label and name_target in label:
                            indices1 = [i for i, x in enumerate(data['source']) \
                                        if x == label.index(name_source)]
                            indices2 = [i for i, x in enumerate(data['target']) \
                                        if x == label.index(name_target)]
                            intersect=set(indices1).intersection(indices2)
                            if len(intersect) > 0:
                                data['value'][intersect.pop()] += carbon_content
                            else:
                                if name_source not in label:
                                    label.append(name_source)
                                    data['source'].append(counter)
                                    node_color.append(self.get_process_color(link[0][0]))
                                    counter += 1
                                else:
                                    label.index(name_source)
                                    data['source'].append(label.index(name_source))


                                if name_target not in label:
                                    label.append(name_target)
                                    data['target'].append(counter)
                                    node_color.append(self.get_process_color(link[1][0]))
                                    counter += 1
                                else:
                                    data['target'].append(label.index(name_target))

                                data['value'].append(carbon_content)

                        else:
                            if name_source not in label:
                                label.append(name_source)
                                data['source'].append(counter)
                                node_color.append(self.get_process_color(link[0][0]))
                                counter += 1
                            else:
                                label.index(name_source)
                                data['source'].append(label.index(name_source))


                            if name_target not in label:
                                label.append(name_target)
                                data['target'].append(counter)
                                node_color.append(self.get_process_color(link[1][0]))
                                counter += 1
                            else:
                                data['target'].append(label.index(name_target))

                            data['value'].append(carbon_content)
                    temp = "rgba(0,0,96,0.2)"
                    data['color'].append(temp)

            else:
                if link[0][1] == 'Material' and link[1][1] == 'Material':
                    if link[0][0] in process_list or link[1][0] in process_list:

                        stream = self.multiplex[link[0]][link[1]][link[2]]
                        carbon_content = stream['carbon_content']*stream['mass_flow_rate']

                        # if carbon_content >= cutoff and link[0][0] not in ignore_list:
                        if link[0][0] not in ignore_list:

                            try:
                                name_source = self.nodes[link[0][0]][process_name]
                            except KeyError:
                                try:
                                    name_source = self.nodes[link[0][0]]['name']
                                except KeyError:
                                    name_source = link[0][0]

                            try:
                                name_target = self.nodes[link[1][0]][process_name]
                            except KeyError:
                                try:
                                    name_target = self.nodes[link[1][0]]['name']
                                except KeyError:
                                    name_target = link[1][0]

                            if name_target == 'Stack':
                                name_target = 'Environment'

                            if name_source in label and name_target in label:
                                indices1 = [i for i, x in enumerate(data['source']) \
                                            if x == label.index(name_source)]
                                indices2 = [i for i, x in enumerate(data['target']) \
                                            if x == label.index(name_target)]
                                intersect=set(indices1).intersection(indices2)
                                if len(intersect) > 0:
                                    data['value'][intersect.pop()] += carbon_content

                                else:
                                    if name_source not in label:
                                        label.append(name_source)
                                        data['source'].append(counter)
                                        node_color.append(self.get_process_color(link[0][0]))
                                        counter += 1
                                    else:
                                        label.index(name_source)
                                        data['source'].append(label.index(name_source))

                                    if name_target not in label:
                                        label.append(name_target)
                                        data['target'].append(counter)
                                        node_color.append(self.get_process_color(link[1][0]))
                                        counter += 1
                                    else:
                                        data['target'].append(label.index(name_target))

                                    data['value'].append(carbon_content)

                            else:

                                if name_source not in label:
                                    label.append(name_source)
                                    data['source'].append(counter)
                                    node_color.append(self.get_process_color(link[0][0]))
                                    counter += 1
                                else:
                                    label.index(name_source)
                                    data['source'].append(label.index(name_source))


                                if name_target not in label:
                                    label.append(name_target)
                                    data['target'].append(counter)
                                    node_color.append(self.get_process_color(link[1][0]))
                                    counter += 1
                                else:
                                    data['target'].append(label.index(name_target))

                                data['value'].append(carbon_content)
                        temp = "rgba(0,0,96,0.2)"
                        data['color'].append(temp)
        data["color"] = []
        for src in data["source"]:
            data["color"].append(node_color[src].replace("0.8","0.55"))
        for i in range(0,len(data["value"]),1):
            if data["value"][i] <= cutoff:
                data["value"][i] = 0
        if bold is True:
            label = ["<b>"+name for name in label]

        fig = go.Figure(data=[go.Sankey(
            valueformat = ".2f",
            valuesuffix = "ktonne/oper-year",
            node = dict(
            pad = fig_pad,
            thickness = fig_thickness,
            line = dict(color = "black", width = 1.5),
            label = label,
            color = node_color
            ),
            link = data
        )])

        if title is True:
            fig.update_layout(title_text="Carbon flows", font_size=text_font,
            font_family = "Times New Roman",
            autosize = False,
            width = fig_width,
            height = fig_height)
        else:
            fig.update_layout(font_size=text_font,
            font_family = "Times New Roman",
            autosize = False,
            width = fig_width,
            height = fig_height)

        config = {
        'toImageButtonOptions': { 'height': None, 'width': None}
        }
        fig.show(config=config)
        fig.write_image("test_carbon_sankey.pdf")

        if print_excel is True:
            # Open Excel
            sheet_name = "Sheet 1"
            work_book = openpyxl.load_workbook("sankey_data.xlsx")

            # if the sheet doesn't exist, create a new sheet
            try:
                sheet = work_book[sheet_name]
                work_book.remove(sheet)
                sheet = work_book.create_sheet(sheet_name)
            except KeyError:
                sheet = work_book.create_sheet(sheet_name)

            # Print degree centrality
            sheet.cell(row = 1, column = 1).value = "Source"
            sheet.cell(row = 1, column = 2).value = "Target"
            sheet.cell(row = 1, column = 3).value = "Mass flow of C [kt/y]"
            sheet.cell(row = 1, column = 4).value = "Color"
            for i in range(0,len(data["source"]),1):
                sheet.cell(row = 2+i, column = 1).value = label[data["source"][i]]
                sheet.cell(row = 2+i, column = 2).value = label[data["target"][i]]
                sheet.cell(row = 2+i, column = 3).value = data["value"][i]
                sheet.cell(row = 2+i, column = 4).value = data["color"][i]

            work_book.save("sankey_data.xlsx")



    def get_process_color(self, uid):
        """Get the process colors based on their subcluster"""
        sub_cluster_colors = {
                "A": "rgba(123,123,123,0.8)",
                "B": "rgba(0,176,80,0.8)",
                "CB": "rgba(216,216,216,0.8)",
                "CL": "rgba(91,155,213,0.8)",
                "E": "rgba(48,84,151,0.8)",
                "M": "rgba(255,192,0,0.8)",
                "N": "rgba(0,127,127,0.8)",
                "O": "rgba(112,48,160,0.8)",
                "P": "rgba(192,0,0,0.8)",
                "U": "rgba(83,129,53,0.8)"
            }
        if uid in self.process_nodes:
            sub_id = re.findall('\\d*\\D+',uid)
            color = sub_cluster_colors[sub_id[0]]
            return color
        else:
            if uid == "MRKT":
                return "rgba(140, 86, 75, 0.8)"
            else:
                return "rgba(127, 127, 127, 0.8)"


    def material_sankey(self, process_list="", ignore_list="", cutoff=0.1, fig_width=1500,\
                        fig_height=750, fig_pad=150, fig_thickness = 10, text_font = 30,\
                        title=True):
        '''Depicts all the carbon flows in the cluster in a sankey diagram'''
        label = []

        data = {}
        data['source'] = []
        data['target'] = []
        data['value'] = []
        data['color'] = []
        data['label'] = []
        counter = 0

        for link in self.multiplex.get_edges():
            if not process_list:
                if link[0][1] == 'Material' and link[1][1] == 'Material':
                    stream = self.multiplex[link[0]][link[1]][link[2]]
                    material_flow = stream['mass_flow_rate']

                    if material_flow >= cutoff and link[0][0] not in ignore_list:
                        try:
                            name_source = self.nodes[link[0][0]]['name']
                        except KeyError:
                            name_source = link[0][0]

                        try:
                            name_target = self.nodes[link[1][0]]['name']
                        except KeyError:
                            name_target = link[1][0]

                        if name_target == 'Stack':
                            name_target = 'Environment'

                        if name_source in label and name_target in label:
                            indices1 = [i for i, x in enumerate(data['source']) \
                                        if x == label.index(name_source)]
                            indices2 = [i for i, x in enumerate(data['target']) \
                                        if x == label.index(name_target)]
                            intersect=set(indices1).intersection(indices2)
                            if len(intersect) > 0:
                                data['value'][intersect.pop()] += material_flow
                            else:
                                if name_source not in label:
                                    label.append(name_source)
                                    data['source'].append(counter)
                                    counter += 1
                                else:
                                    label.index(name_source)
                                    data['source'].append(label.index(name_source))


                                if name_target not in label:
                                    label.append(name_target)
                                    data['target'].append(counter)
                                    counter += 1
                                else:
                                    data['target'].append(label.index(name_target))

                                data['value'].append(material_flow)

                        else:
                            if name_source not in label:
                                label.append(name_source)
                                data['source'].append(counter)
                                counter += 1
                            else:
                                label.index(name_source)
                                data['source'].append(label.index(name_source))


                            if name_target not in label:
                                label.append(name_target)
                                data['target'].append(counter)
                                counter += 1
                            else:
                                data['target'].append(label.index(name_target))

                            data['value'].append(material_flow)
                    color = list(np.random.choice(range(256), size=3))
                    temp = f"rgba({color[0]},{color[1]},{color[2]},0.8)"
                    data['color'].append(temp)

            else:
                if link[0][1] == 'Material' and link[1][1] == 'Material':
                    if link[0][0] in process_list or link[1][0] in process_list:

                        stream = self.multiplex[link[0]][link[1]][link[2]]
                        material_flow = stream['mass_flow_rate']

                        if material_flow >= cutoff and link[0][0] not in ignore_list:

                            try:
                                name_source = self.nodes[link[0][0]]['name']
                            except KeyError:
                                name_source = link[0][0]

                            try:
                                name_target = self.nodes[link[1][0]]['name']
                            except KeyError:
                                name_target = link[1][0]

                            if name_target == 'Stack':
                                name_target = 'Environment'

                            if name_source in label and name_target in label:
                                indices1 = [i for i, x in enumerate(data['source']) \
                                            if x == label.index(name_source)]
                                indices2 = [i for i, x in enumerate(data['target']) \
                                            if x == label.index(name_target)]
                                intersect=set(indices1).intersection(indices2)
                                if len(intersect) > 0:
                                    data['value'][intersect.pop()] += material_flow

                                else:
                                    if name_source not in label:
                                        label.append(name_source)
                                        data['source'].append(counter)
                                        counter += 1
                                    else:
                                        label.index(name_source)
                                        data['source'].append(label.index(name_source))

                                    if name_target not in label:
                                        label.append(name_target)
                                        data['target'].append(counter)
                                        counter += 1
                                    else:
                                        data['target'].append(label.index(name_target))

                                    data['value'].append(material_flow)

                            else:

                                if name_source not in label:
                                    label.append(name_source)
                                    data['source'].append(counter)
                                    counter += 1
                                else:
                                    label.index(name_source)
                                    data['source'].append(label.index(name_source))


                                if name_target not in label:
                                    label.append(name_target)
                                    data['target'].append(counter)
                                    counter += 1
                                else:
                                    data['target'].append(label.index(name_target))

                                data['value'].append(material_flow)
                        color = list(np.random.choice(range(256), size=3))
                        temp = f"rgba({color[0]},{color[1]},{color[2]},0.8)"
                        data['color'].append(temp)

        fig = go.Figure(data=[go.Sankey(
            valueformat = ".2f",
            valuesuffix = "ktonne/oper-year",
            node = dict(
            pad = fig_pad,
            thickness = fig_thickness,
            line = dict(color = "black", width = 0.5),
            label = label,
            color = "blue"
            ),
            link = data
        )])

        if title is True:
            fig.update_layout(title_text="Material flows", font_size=text_font,
            autosize = False,
            width = fig_width,
            height = fig_height)
        else:
            fig.update_layout(font_size=text_font,
            autosize = False,
            width = fig_width,
            height = fig_height)

        config = {
        'toImageButtonOptions': { 'height': None, 'width': None}
        }
        fig.show(config=config)



    def water_sankey(self, process_list ="",ignore_list ="" ,cutoff = 0.1, fig_width=1500, \
                    fig_height=750, fig_pad=150, fig_thickness = 15, text_font=30):
        '''Depicts all the water flows in the cluster in a sankey diagram'''
        label = []

        data = {}
        data['source'] = []
        data['target'] = []
        data['value'] = []
        data['color'] = []
        data['label'] = []
        counter = 0

        for link in self.multiplex.get_edges():
            if not process_list:
                if link[0][1] == 'Material' and link[1][1] == 'Material':

                    stream = self.multiplex[link[0]][link[1]][link[2]]
                    try:
                        water = stream['mass_fraction']['H2O']*stream['mass_flow_rate']
                    except KeyError:
                        water = 0

                    if water >= cutoff and link[0][0] not in ignore_list:
                        try:
                            name_source = self.nodes[link[0][0]]['name']
                        except KeyError:
                            name_source = link[0][0]

                        try:
                            name_target = self.nodes[link[1][0]]['name']
                        except KeyError:
                            name_target = link[1][0]

                        if name_target == 'Stack':
                            name_target = 'Environment'

                        if name_source in label and name_target in label:
                            indices1 = [i for i, x in enumerate(data['source']) \
                                        if x == label.index(name_source)]
                            indices2 = [i for i, x in enumerate(data['target']) \
                                        if x == label.index(name_target)]
                            intersect=set(indices1).intersection(indices2)
                            if len(intersect) > 0:
                                data['value'][intersect.pop()] += water
                            else:
                                if name_source not in label:
                                    label.append(name_source)
                                    data['source'].append(counter)
                                    counter += 1
                                else:
                                    label.index(name_source)
                                    data['source'].append(label.index(name_source))

                                if name_target not in label:
                                    label.append(name_target)
                                    data['target'].append(counter)
                                    counter += 1
                                else:
                                    data['target'].append(label.index(name_target))

                                data['value'].append(water)
                        else:
                            if name_source not in label:
                                label.append(name_source)
                                data['source'].append(counter)
                                counter += 1
                            else:
                                label.index(name_source)
                                data['source'].append(label.index(name_source))

                            if name_target not in label:
                                label.append(name_target)
                                data['target'].append(counter)
                                counter += 1
                            else:
                                data['target'].append(label.index(name_target))
                            data['value'].append(water)
                        color = list(np.random.choice(range(256), size=3))
                        temp = f"rgba({color[0]},{color[1]},{color[2]},0.8)"
                        data['color'].append(temp)

            else:
                if link[0][1] == 'Material' and link[1][1] == 'Material':
                    if link[0][0] in process_list or link[1][0] in process_list:

                        stream = self.multiplex[link[0]][link[1]][link[2]]
                        try:
                            water = stream['mass_fraction']['H2O']*stream['mass_flow_rate']
                        except KeyError:
                            water = 0

                        if water >= cutoff and link[0][0] not in ignore_list:

                            try:
                                name_source = self.nodes[link[0][0]]['name']
                            except KeyError:
                                name_source = link[0][0]

                            try:
                                name_target = self.nodes[link[1][0]]['name']
                            except KeyError:
                                name_target = link[1][0]

                            if name_target == 'Stack':
                                name_target = 'Environment'

                            if name_source in label and name_target in label:
                                indices1 = [i for i, x in enumerate(data['source']) \
                                            if x == label.index(name_source)]
                                indices2 = [i for i, x in enumerate(data['target']) \
                                            if x == label.index(name_target)]
                                intersect=set(indices1).intersection(indices2)
                                if len(intersect) > 0:
                                    data['value'][intersect.pop()] += water
                                else:
                                    if name_source not in label:
                                        label.append(name_source)
                                        data['source'].append(counter)
                                        counter += 1
                                    else:
                                        label.index(name_source)
                                        data['source'].append(label.index(name_source))

                                    if name_target not in label:
                                        label.append(name_target)
                                        data['target'].append(counter)
                                        counter += 1
                                    else:
                                        data['target'].append(label.index(name_target))

                                        data['value'].append(water)

                            else:
                                if name_source not in label:
                                    label.append(name_source)
                                    data['source'].append(counter)
                                    counter += 1
                                else:
                                    label.index(name_source)
                                    data['source'].append(label.index(name_source))

                                if name_target not in label:
                                    label.append(name_target)
                                    data['target'].append(counter)
                                    counter += 1
                                else:
                                    data['target'].append(label.index(name_target))

                                data['value'].append(water)
                            color = list(np.random.choice(range(256), size=3))

                            temp = f"rgba({color[0]},{color[1]},{color[2]},0.8)"
                            data['color'].append(temp)

        fig = go.Figure(data=[go.Sankey(
            valueformat = ".2f",
            valuesuffix = "ktonne/oper-year",
            node = dict(
            pad = fig_pad,
            thickness = fig_thickness,
            line = dict(color = "black", width = 0.5),
            label = label,
            color = "blue"
            ),
            link = data
        )])

        fig.update_layout(title_text="Water flows", font_size=text_font,
        autosize = False,
        width = fig_width,
        height = fig_height)
        fig.show()


    def steam_sankey(self, process_list="", ignore_list="", cutoff=0.1, fig_width=1500, \
                    fig_height=750, fig_pad=150, fig_thickness = 10, text_font = 30):
        '''Depicts all the carbon flows in the cluster in a sankey diagram'''
        label = []

        data = {}
        data['source'] = []
        data['target'] = []
        data['value'] = []
        data['color'] = []
        data['label'] = []
        counter = 0

        for link in self.multiplex.get_edges():
            if not process_list:
                if link[0][1] == 'Steam' and link[1][1] == 'Steam':
                    stream = self.multiplex[link[0]][link[1]][link[2]]
                    steam_amount = stream['energy']

                    if steam_amount >= cutoff and link[0][0] not in ignore_list:
                        try:
                            name_source = self.nodes[link[0][0]]['name']
                        except KeyError:
                            name_source = link[0][0]

                        try:
                            name_target = self.nodes[link[1][0]]['name']
                        except KeyError:
                            name_target = link[1][0]

                        if name_target == 'Stack':
                            name_target = 'Environment'

                        if name_source in label and name_target in label:
                            indices1 = [i for i, x in enumerate(data['source']) \
                                        if x == label.index(name_source)]
                            indices2 = [i for i, x in enumerate(data['target']) \
                                        if x == label.index(name_target)]
                            intersect=set(indices1).intersection(indices2)
                            if len(intersect) > 0:
                                data['value'][intersect.pop()] += steam_amount
                            else:
                                if name_source not in label:
                                    label.append(name_source)
                                    data['source'].append(counter)
                                    counter += 1
                                else:
                                    label.index(name_source)
                                    data['source'].append(label.index(name_source))


                                if name_target not in label:
                                    label.append(name_target)
                                    data['target'].append(counter)
                                    counter += 1
                                else:
                                    data['target'].append(label.index(name_target))

                                data['value'].append(steam_amount)

                        else:
                            if name_source not in label:
                                label.append(name_source)
                                data['source'].append(counter)
                                counter += 1
                            else:
                                label.index(name_source)
                                data['source'].append(label.index(name_source))


                            if name_target not in label:
                                label.append(name_target)
                                data['target'].append(counter)
                                counter += 1
                            else:
                                data['target'].append(label.index(name_target))

                            data['value'].append(steam_amount)
                    color = list(np.random.choice(range(256), size=3))
                    temp = f"rgba({color[0]},{color[1]},{color[2]},0.8)"
                    data['color'].append(temp)

            else:
                if link[0][1] == 'Steam' and link[1][1] == 'Steam':
                    if link[0][0] in process_list or link[1][0] in process_list:

                        stream = self.multiplex[link[0]][link[1]][link[2]]
                        steam_amount = stream['energy']

                        if steam_amount >= cutoff and link[0][0] not in ignore_list:

                            try:
                                name_source = self.nodes[link[0][0]]['name']
                            except KeyError:
                                name_source = link[0][0]

                            try:
                                name_target = self.nodes[link[1][0]]['name']
                            except KeyError:
                                name_target = link[1][0]

                            if name_target == 'Stack':
                                name_target = 'Environment'

                            if name_source in label and name_target in label:
                                indices1 = [i for i, x in enumerate(data['source']) \
                                            if x == label.index(name_source)]
                                indices2 = [i for i, x in enumerate(data['target']) \
                                            if x == label.index(name_target)]
                                intersect=set(indices1).intersection(indices2)
                                if len(intersect) > 0:
                                    data['value'][intersect.pop()] += steam_amount

                                else:
                                    if name_source not in label:
                                        label.append(name_source)
                                        data['source'].append(counter)
                                        counter += 1
                                    else:
                                        label.index(name_source)
                                        data['source'].append(label.index(name_source))

                                    if name_target not in label:
                                        label.append(name_target)
                                        data['target'].append(counter)
                                        counter += 1
                                    else:
                                        data['target'].append(label.index(name_target))

                                    data['value'].append(steam_amount)

                            else:

                                if name_source not in label:
                                    label.append(name_source)
                                    data['source'].append(counter)
                                    counter += 1
                                else:
                                    label.index(name_source)
                                    data['source'].append(label.index(name_source))


                                if name_target not in label:
                                    label.append(name_target)
                                    data['target'].append(counter)
                                    counter += 1
                                else:
                                    data['target'].append(label.index(name_target))

                                data['value'].append(steam_amount)
                        color = list(np.random.choice(range(256), size=3))
                        temp = f"rgba({color[0]},{color[1]},{color[2]},0.8)"
                        data['color'].append(temp)

        fig = go.Figure(data=[go.Sankey(
            valueformat = ".2f",
            valuesuffix = "TJ/oper-year",
            node = dict(
            pad = fig_pad,
            thickness = fig_thickness,
            line = dict(color = "black", width = 0.5),
            label = label,
            color = "blue"
            ),
            link = data
        )])

        fig.update_layout(title_text="Steam energy flows", font_size=text_font,
        autosize = False,
        width = fig_width,
        height = fig_height)

        config = {
        'toImageButtonOptions': { 'height': None, 'width': None}
        }
        fig.show(config=config)


    def carbon_process(self, process):
        '''
        Calculate the total amount of carbon entering the process,
        the total amount of carbon in the waste streams and
        the total amount of carbon in the product stream.
        [ktonne / operating-year]
        '''
        carbon_in = 0
        carbon_waste = 0
        carbon_prod = 0

        for link in self.multiplex.get_edges():
            if link[0][0] == process:
                if link[1][0] == 'PROD' or link[1][0] in self.process_nodes.keys():
                    try:
                        stream = self.multiplex[link[0]][link[1]][link[2]]
                        carbon = stream['mass_flow_rate'] * stream['carbon_content']
                        carbon_prod += carbon
                    except KeyError:
                        pass
                else:
                    try:
                        stream = self.multiplex[link[0]][link[1]][link[2]]
                        carbon = stream['mass_flow_rate'] * stream['carbon_content']
                        carbon_waste += carbon
                    except KeyError:
                        pass
            if link[1][0] == process:
                try:
                    stream = self.multiplex[link[0]][link[1]][link[2]]
                    carbon = stream['mass_flow_rate'] * stream['carbon_content']
                    carbon_in += carbon
                except KeyError:
                    pass

        return carbon_in, carbon_waste, carbon_prod


    def carbon_cluster(self):
        '''
        Calculate the total amount of carbon entering the cluster,
        the total amount of carbon in the cluster waste streams and
        the total amount of carbon in the cluster product streams
        [ktonne / operating-year]
        '''
        carbon_in = 0
        carbon_envi = 0
        carbon_prod = 0

        for link in self.multiplex.get_edges(data=True):
            if link[0][1] == "Material" and link[1][1] == "Material":
                if link[0][0] == "MRKT":
                    if link[1][0] in self.process_nodes.keys():
                        for stream in link[3]:
                            carbon_in += stream["carbon_flow_rate"]
                elif link[1][0] == "PROD":
                    for stream in link[3]:
                        print(stream)
                        carbon_prod += stream["carbon_flow_rate"]
                elif link[1][0] == "ENVI":
                    for stream in link[3]:
                        carbon_envi += stream["carbon_flow_rate"]

        return carbon_in, carbon_envi, carbon_prod, carbon_prod/carbon_in


    def water_process_old(self, process):
        '''
        Returns the total amount of water entering the process,
        total water to the chlorinated waste treatment,
        total water to waste water treatment,
        total water in the product streams,
        total water in the air,
        total water in the source.
        [ktonne / operating-year]
        '''

        water_in = 0
        water_lq_chlor = 0
        water_wwt = 0
        water_prod = 0
        water_air = 0
        water_source = 0

        for link in self.multiplex.get_edges():
            if link[0][0] == process:
                stream = self.multiplex[link[0]][link[1]][link[2]]
                if link[1][0] == 'PROD' or link[1][0] in self.process_nodes.keys():
                    try:
                        water = stream['mass_flow_rate'] * stream['mass_fraction']['H2O']
                        water_prod += water
                    except KeyError:
                        pass
                elif link[1][0] == 'LQCL':
                    try:
                        water = stream['mass_fraction']['H2O'] * stream['mass_flow_rate']
                        water_lq_chlor += water
                    except KeyError:
                        pass
                elif link[1][0] == 'WWT':
                    try:
                        water = stream['mass_fraction']['H2O'] * stream['mass_flow_rate']
                        water_wwt += water
                    except KeyError:
                        pass
                else:
                    if stream['vapor_fraction'] == 1:
                        try:
                            water = stream['mass_fraction']['H2O'] * stream['mass_flow_rate']
                            water_air += water
                        except KeyError:
                            pass
                    else:
                        try:
                            water = stream['mass_fraction']['H2O'] * stream['mass_flow_rate']
                            water_source += water
                        except KeyError:
                            pass
            if link[1][0] == process:
                try:
                    stream = self.multiplex[link[0]][link[1]][link[2]]
                    water = stream['mass_flow_rate'] * stream['mass_fraction']['H2O']
                    water_in += water
                except KeyError:
                    pass

        return water_in, water_lq_chlor, water_wwt, water_prod, water_air, water_source


    def water_process(self, process):
        '''
        Returns the total amount of water entering the process,
        total water to the chlorinated waste treatment,
        total water to waste water treatment,
        total water in the product streams,
        total water in the air,
        total water in the source.
        [ktonne / operating-year]
        '''
        water_in = 0
        water_chlor = 0
        water_wwt = 0
        water_prod = 0
        water_hct = 0


        for link in self.multiplex.get_edges():
            if link[0][0] == process:
                stream = self.multiplex[link[0]][link[1]][link[2]]
                if link[1][0] == 'PROD' or link[1][0] in self.process_nodes.keys():
                    try:
                        water = stream['mass_flow_rate'] * stream['mass_fraction']['H2O']
                        water_prod += water
                    except KeyError:
                        pass
                elif link[1][0] == 'LQCL':
                    try:
                        water = stream['mass_fraction']['H2O'] * stream['mass_flow_rate']
                        water_chlor += water
                    except KeyError:
                        pass
                elif link[1][0] == 'WWT':
                    try:
                        water = stream['mass_fraction']['H2O'] * stream['mass_flow_rate']
                        water_wwt += water
                    except KeyError:
                        pass
                elif link[1][0] == 'HCT':
                    try:
                        water = stream['mass_fraction']['H2O'] * stream['mass_flow_rate']
                        water_hct += water
                    except KeyError:
                        pass

            if link[1][0] == process:
                try:
                    stream = self.multiplex[link[0]][link[1]][link[2]]
                    water = stream['mass_flow_rate'] * stream['mass_fraction']['H2O']
                    water_in += water
                except KeyError:
                    pass

        return water_in, water_chlor, water_wwt, water_prod, water_hct


    def water_cluster(self):
        '''
        Returns the total amount of water entering the cluster,
        total water to the chlorinated waste treatment,
        total water to waste water treatment,
        total water in the product streams,
        total water in the air,
        total water in the source
        [ktonne / operating-year]
        '''
        water_in = 0
        water_lq_chlor = 0
        water_wwt = 0
        water_prod = 0
        water_air = 0
        water_source = 0

        tags,layers,_link_dict=  self.multiplex.get_layers()
        layer_in = tags.index('IN')
        layer_out = tags.index('OUT')

        for link in layers[layer_in].edges:
            stream = self.multiplex[link[0]][link[1]][link[2]]
            try:
                water = stream['mass_fraction']['H2O'] * stream['mass_flow_rate']
                water_in += water
            except KeyError:
                pass

        for link in layers[layer_out].edges:
            stream = self.multiplex[link[0]][link[1]][link[2]]
            if link[1][0] == 'PROD':
                try:
                    water = stream['mass_fraction']['H2O'] * stream['mass_flow_rate']
                    water_prod += water
                except KeyError:
                    pass
            elif link[1][0] == 'LQCL':
                try:
                    water = stream['mass_fraction']['H2O'] * stream['mass_flow_rate']
                    water_lq_chlor += water
                except KeyError:
                    pass
            elif link[1][0] == 'WWT':
                try:
                    water = stream['mass_fraction']['H2O'] * stream['mass_flow_rate']
                    water_wwt += water
                except KeyError:
                    pass
            else:
                if stream['vapor_fraction'] == 1:
                    try:
                        water = stream['mass_fraction']['H2O'] * stream['mass_flow_rate']
                        water_air += water
                    except KeyError:
                        pass
                else:
                    try:
                        water = stream['mass_fraction']['H2O'] * stream['mass_flow_rate']
                        water_source += water
                    except KeyError:
                        pass

        return water_in, water_lq_chlor, water_wwt, water_prod, water_air, water_source


    def co2_process(self, process):
        '''
        Calculates the amount of direct CO2 emissions of a process
        [ktonne / operating-year]
        '''

        co2_out = 0

        for link in self.multiplex.get_edges():
            if link[0][0] == process and link[1][0] == 'ENVI':
                try:
                    stream = self.multiplex[link[0]][link[1]][link[2]]
                    co2 = stream['mass_flow_rate'] * stream['mass_fraction']['CO2']
                    co2_out += co2
                except KeyError:
                    pass

            elif link[0][0] == process and link[1][0] == 'STCK':
                stream = self.multiplex[link[0]][link[1]][link[2]]
                co2 = stream['carbon_content']*stream['mass_flow_rate']/12.011*44.0095
                co2_out += co2

        return co2_out


    def co2_cluster(self):
        '''
        Calculates the amount of direct CO2 emissions of the cluster
        [ktonne / operating-year]
        '''
        co2_out = 0

        tags,layers,_link_dict=  self.multiplex.get_layers()
        layer_out = tags.index('OUT')

        for link in layers[layer_out].edges:
            data = self.multiplex[link[0]][link[1]][link[2]]

            if link[1][0] == 'ENVI:':
                try:
                    co2 = data['mass_fraction']['CO2'] * data['mass_flow_rate']
                    co2_out += co2
                except KeyError:
                    co2_out += 0
            elif link[1][0] == 'STCK':
                co2 = data['carbon_content']*data['mass_flow_rate']/12.011*44.0095
                co2_out += co2

        return co2_out


    def energy_process(self, process, ignore_list=''):
        '''Retrieves the energy consumption of a process and returns the values as a piechart'''

        label_dict = {
        'LLPS': 'Very low pressure steam<br>(T=142,7°C)',
        'LPS' : 'Low pressure steam<br>(T=155,5°C)',
        'MPS' : 'Medium pressure steam<br>(T=214,9°C)',
        'HPS' : 'High pressure steam<br>(T=265,2°C)',
        'HHPS': 'Very high pressure steam>br> (T=311°C)',
        'Electricity': 'Electricity',
        'NG': 'Natural gas',
        'R134A': 'Refrigerant<br>R134A',
        'R717': 'Refrigerant<br>R717',
        'R410A': 'Refrigerant<br>R410A',
        'R41': 'Refrigerant<br>R41',
        'R1150': "Refrigerant<br>R1150",
        'Cooling water': 'Cooling water',
        'Chilled water': 'Chilled water'
        }

        color_dict = {
            'LLPS': 'bluegreen',
            'LPS':  'blue',
            'MPS':  'blueviolet',
            'HPS':  'red',
            'HHPS': 'darkred',
            'Electricity': 'orange',
            'NG':   'yellowgreen',
            'R134A': 'pink',
            'R717': 'black',
            'R410A': 'brown',
            'R41': 'turquoise',
            'R1150': "red",
            'Cooling water': 'lightblue',
            'Chilled water': 'darkblue'
        }

        energy = self.nodes[process]['energy_consumption']
        values=[round(x,0) for name,x in energy.items() if name not in ignore_list and x != 0]
        utilities=[name for name,x in energy.items() if name not in ignore_list and  x != 0]

        labels = [label_dict[name] for name in utilities]
        colors = [color_dict[name] for name in utilities]
        process_name = self.nodes[process]['name']

        fig = go.Figure(data=[go.Pie(labels=labels,
                             values=values)])
        fig.update_traces(hoverinfo='label+percent', textinfo='value', textfont_size=30,
                        marker=dict(colors=colors, line=dict(color='#000000', width=2)))
        fig.update_layout(title_text = f'{process}. {process_name}<br>net energy consumption \
                        in TJ/year',
                        autosize=False,
                        width=1000,
                        height=1000,
                        font=dict(size=30),
                        legend_font=dict(size=25))

        config = {
        'toImageButtonOptions': { 'height': None, 'width': None, }
        }
        fig.show(config=config)


    def energy_cluster(self, ignore_list='' ,normalized=False, title=True):
        '''Retrieves the overall energy consumption of the cluster and
        presents the values in a piechart'''

        label_dict = {
        'LLPS': 'Very low pressure steam<br>(T=142,7°C)',
        'LPS' : 'Low pressure steam<br>(T=155,5°C)',
        'MPS' : 'Medium pressure steam<br>(T=214,9°C)',
        'HPS' : 'High pressure steam<br>(T=265,2°C)',
        'HHPS': 'Very high pressure steam<br>(T=311°C)',
        'Electricity': 'Electricity',
        'NG': 'Natural gas',
        'R134A': 'Refrigerant<br>R134A',
        'R717': 'Refrigerant<br>R717',
        'R410A': 'Refrigerant<br>R410A',
        'R41': 'Refrigerant<br>R41',
        'R1150': "Refrigerant<br>R1150",
        'Cooling water': 'Cooling water',
        'Chilled water': 'Chilled water'
        }

        color_dict = {
            'LLPS': 'bluegreen',
            'LPS':  'blue',
            'MPS':  'blueviolet',
            'HPS':  'red',
            'HHPS': 'darkred',
            'Electricity': 'orange',
            'NG':   'yellowgreen',
            'R134A': 'pink',
            'R717': 'black',
            'R410A': 'brown',
            'R41': 'turquoise',
            'R1150': "red",
            'Cooling water': 'lightblue',
            'Chilled water': 'darkblue'
        }
        if normalized is True:
            carbon_prod = 0
            tags,layers,_link_dict=  self.multiplex.get_layers()
            layer_out = tags.index('OUT')

            for link in layers[layer_out].edges:
                if link[1][0] == 'PROD':
                    data = self.multiplex[link[0]][link[1]][link[2]]
                    carbon_prod += data['carbon_content'] * data['mass_flow_rate']

            energy_total = Counter({})
            for node in self.process_nodes.values():
                energy_total += Counter(node['energy_consumption'])

            values=[round(x/carbon_prod,1) for name,x in energy_total.items() \
                    if name not in ignore_list and x != 0]
            utilities=[name for name,x in energy_total.items() \
                    if name not in ignore_list and x != 0]
            figure_title = "Cluster net energy intensity<br>in TJ/ktonne of carbon product"

        else:
            energy_total = Counter({})
            for node in self.process_nodes.values():
                energy_total += Counter(node['energy_consumption'])

            values=[round(x,0) for name,x in energy_total.items() \
                    if name not in ignore_list and x != 0]
            utilities=[name for name,x in energy_total.items() \
                    if name not in ignore_list and  x != 0]
            figure_title = "Cluster<br>net energy consumption in TJ/year"

        labels = [label_dict[name] for name in utilities]
        colors = [color_dict[name] for name in utilities]

        fig = go.Figure(data=[go.Pie(labels=labels,
                        values=values)])
        fig.update_traces(hoverinfo='label+percent', textinfo='value+percent', textfont_size=25,
                        marker=dict(colors=colors, line=dict(color='#000000', width=2)))
        if title is True:
            fig.update_layout(title_text = figure_title,
            autosize = False,
            width=1000,
            height=1000,
            font=dict(size=25),
            legend_font=dict(size=25))
        else:
            fig.update_layout(autosize = False,
            width=1000,
            height=1000,
            font=dict(size=25),
            legend_font=dict(size=25))

        config = {
        'toImageButtonOptions': { 'height': None, 'width': None, }
        }
        fig.show(config=config)
        #"""fig.write_image("figures/energy_process.svg")"""


    def scatter_steam(self, ignore_list="", height=1000, width =1000, font_size=10, \
                      print_excel=False):
        """Method for plotting the steam intensity over the steam consumption in a scatter plot"""
        steam_consumption = {}
        steam_intensity = {}

        for node, node_value in self.process_nodes.items():
            process_energy = 0
            for name, value in node_value['energy_consumption'].items():
                if value <= 0:
                    continue
                if name == 'LLPS':
                    process_energy += value
                if name == 'LPS':
                    process_energy += value
                if name == 'MPS':
                    process_energy += value
                if name == 'HPS':
                    process_energy += value
                if name == "HHPS":
                    process_energy += value

            process_carbon = 0

            for link in self.multiplex[node,'Material']:
                if link[0] == node:
                    continue
                if "U" in link[0]:
                    # Ignore the carbon in streams send to utility units that is
                    # used as energy source
                    continue
                if link[0] in self.process_nodes.keys():
                    for stream in self.multiplex[node,'Material'][link].values():
                        process_carbon += stream['carbon_flow_rate']
                if link[0] == 'PROD':
                    for stream in self.multiplex[node,'Material'][link].values():
                        process_carbon += stream['carbon_flow_rate']

            if process_energy >= 0.1:
                steam_consumption[node] = process_energy

            try:
                if process_carbon >= 0.1:
                    steam_intensity[node] = process_energy/process_carbon
            except ZeroDivisionError:
                pass

        # Remove processes that are not in both the steam consumption and steam intensity lists
        diff = set(steam_intensity) - set(steam_consumption)
        for x in diff:
            steam_intensity.pop(x)
        diff = set(steam_consumption) - set(steam_intensity)
        for x in diff:
            steam_consumption.pop(x)

        for process in ignore_list:
            steam_consumption.pop(process)
            steam_intensity.pop(process)
        # x and y given as array_like objects
        text = [x for x in steam_consumption]

        fig = px.scatter(x=steam_consumption,y=steam_intensity,text=text, log_y=True, log_x=True,\
                        labels={'x':'Steam consumption<br>TJ/year',\
                        'y':'Steam intensity <br>[TJ/(ktonne of carbon X year) ]'},\
                        width=width, height=height)
        fig.update_traces(textposition='top center', marker={'size':10})
        fig.update_layout(font_family="Time New Roman", font_size=font_size)
        fig.show()
        fig.write_image("scatter_steam.pdf")

        if print_excel is True:
            # Open Excel
            sheet_name = "Sheet 1"
            work_book = openpyxl.load_workbook("scatter_steam_data.xlsx")

            # if the sheet doesn't exist, create a new sheet
            try:
                sheet = work_book[sheet_name]
                work_book.remove(sheet)
                sheet = work_book.create_sheet(sheet_name)
            except KeyError:
                sheet = work_book.create_sheet(sheet_name)

            i = 0
            # Print degree centrality
            sheet.cell(row = 1, column = 1).value = "Node"
            sheet.cell(row = 1, column = 2).value = "Steam consumption [TJ/y]"
            sheet.cell(row = 1, column = 3).value = "Steam intensity [TJ/(tonne of C * y)]"
            for key, value in steam_consumption.items():
                sheet.cell(row = 2+i, column=1).value = key
                sheet.cell(row = 2+i, column=2).value = value
                sheet.cell(row = 2+i, column=3).value = steam_intensity[key]
                i+=1

            work_book.save("scatter_steam_data.xlsx")


    def scatter_co2(self, ignore_list="", height=1000, width=1000, \
                      font_size=10, print_excel=False):

        """Method for plotting the CO2 intensity over the CO2 emissions in a scatter plot"""

        co2_emission = {}
        co2_intensity = {}

        co2_emission_util = {}
        co2_intensity_util = {}

        for node in self.process_nodes.keys():
            if "U" not in node or node == "U1":
                process_carbon = 0
                for link in self.multiplex[node,'Material']:
                    if link[0] == node:
                        continue
                    if "U" in link[0]:
                        # Ignore the carbon in streams send to utility units that
                        # is used as energy source
                        continue
                    if link[0] in self.process_nodes.keys():
                        for stream in self.multiplex[node,'Material'][link].values():
                            process_carbon += stream['carbon_flow_rate']
                    if link[0] == 'PROD':
                        for stream in self.multiplex[node,'Material'][link].values():
                            process_carbon += stream['carbon_flow_rate']

                emission = self.co2_process(node)
                if emission >= 0.1:
                    co2_emission[node] = emission
                try:
                    if process_carbon >= 0.1:
                        co2_intensity[node] = emission/process_carbon
                except ZeroDivisionError:
                    pass

            elif "U3" in node or "U5" in node or "U6" in node or "U7" in node \
                or "U8" in node or "U9" in node:
                if "U3" in node:
                    util_name = "U3"
                elif "U5" in node:
                    util_name = "U5"
                elif "U6" in node:
                    util_name = "U6"
                elif "U7" in node:
                    util_name = "U7"
                elif "U8" in node:
                    util_name = "U8"
                else:
                    util_name = "U9"

                energy = 0
                for _steam_type, value in self.process_nodes[node]["energy_production"].items():
                    energy += value[0]
                emission = self.co2_process(node)

                if util_name in co2_emission_util:
                    co2_emission_util[util_name] += emission
                else:
                    co2_emission_util[util_name] = emission

                if util_name in co2_intensity_util:
                    try:
                        co2_intensity_util[util_name] += emission/energy
                    except ZeroDivisionError:
                        pass
                else:
                    try:
                        co2_intensity_util[util_name] = emission/energy
                    except ZeroDivisionError:
                        pass
            else:
                emission = self.co2_process(node)

        # Remove processes that are not in both the steam consumption and steam intensity lists
        diff = set(co2_intensity) - set(co2_emission)
        for x in diff:
            co2_intensity.pop(x)

        diff = set(co2_emission) - set(co2_intensity)
        for x in diff:
            co2_emission.pop(x)

        for process in ignore_list:
            co2_emission.pop(process)
            co2_intensity.pop(process)
        # x and y given as array_like objects
        text1 = [x for x in co2_emission]

        diff = set(co2_intensity_util) - set(co2_emission_util)
        for x in diff:
            co2_intensity_util.pop(x)

        diff = set(co2_emission_util) - set(co2_intensity_util)
        for x in diff:
            co2_emission_util.pop(x)

        for process in ignore_list:
            co2_emission_util.pop(process)
            co2_intensity_util.pop(process)
        # x and y given as array_like objects
        text2 = [x for x in co2_emission_util]

        co2_emission = [x for x in co2_emission.values()]
        co2_intensity = [x for x in co2_intensity.values()]
        co2_emission_util = [x for x in co2_emission_util.values()]
        co2_intensity_util = [x for x in co2_intensity_util.values()]

        fig = make_subplots(specs=[[{"secondary_y": True}]])
        fig.add_trace(go.Scatter(x=co2_emission, y=co2_intensity, text=text1, mode="markers+text",\
                                textposition="top center", marker_size=10, \
                                name= "Chemical processes"),secondary_y=False)
        fig.add_trace(go.Scatter(x=co2_emission_util, y=co2_intensity_util, text=text2, \
                                mode="markers+text",textposition="top center", marker_size=10, \
                                name="CHPs and boilers", marker_symbol="diamond"),secondary_y=True)
        fig.update_layout(font_family = "Times New Roman", height=height, width=width,\
                        font_size=font_size,xaxis_title="CO2 emission <br> [ktonne of CO2 / year]",\
                        yaxis_title="CO2 intensity<br>[ktonne of CO2 / ktonne of carbon product]",\
                        yaxis2_title="CO2 intensity<br>[ktonne of CO2 / TJ of steam]",
                        legend=dict(
                        yanchor="top",
                        y=0.99,
                        xanchor="left",
                        x=0.01
                        ))
        fig.update_xaxes(type="log")
        fig.update_yaxes(type="log")
        fig.write_image("scatter_co2_2.pdf")
        fig.show()

        if print_excel is True:
            # Open Excel
            sheet_name = "Sheet 1"
            work_book = openpyxl.load_workbook("scatter_co2_data.xlsx")

            # if the sheet doesn't exist, create a new sheet
            try:
                sheet = work_book[sheet_name]
                work_book.remove(sheet)
                sheet = work_book.create_sheet(sheet_name)
            except KeyError:
                sheet = work_book.create_sheet(sheet_name)

            # Print co2 emissions and intensity of chemical processes
            sheet.cell(row = 1, column = 1).value = "Node"
            sheet.cell(row = 1, column = 2).value = "CO2 emission [kt CO2/y]"
            sheet.cell(row = 1, column = 3).value = "CO2 intensity [kt CO2/(kt C * y)]"
            for i in range(0,len(co2_emission),1):
                sheet.cell(row = 2+i, column=1).value = text1[i]
                sheet.cell(row = 2+i, column=2).value = co2_emission[i]
                sheet.cell(row = 2+i, column=3).value = co2_intensity[i]

            # Print co2 emissions and intensity of utility processes
            sheet.cell(row = 1, column = 6).value = "Node"
            sheet.cell(row = 1, column = 7).value = "CO2 emission [kt/y]"
            sheet.cell(row = 1, column = 8).value = "CO2 intensity [kt/(TJ steam * y)]"
            for j in range(0,len(co2_emission_util),1):
                sheet.cell(row = 2+j, column=6).value = text2[j]
                sheet.cell(row = 2+j, column=7).value = co2_emission_util[j]
                sheet.cell(row = 2+j, column=8).value = co2_intensity_util[j]

            work_book.save("scatter_co2_data.xlsx")


    def scatter_water(self, ignore_list='', height=1000, width=1000, font_size=10):
        """Create a scatter plot of the water consumption of the cluster"""
        water_consumption = {}
        water_intensity = {}
        water_use = {}
        for node in self.process_nodes.keys():
            if "U" in node:
                continue
            else:
                process_carbon = 0
                for link in self.multiplex[node,'Material']:
                    if link[0] == node:
                        continue
                    if "U" in link[0]:
                        # Ignore the carbon in streams send to utility units that
                        # is used as energy source
                        continue
                    if link[0] in self.process_nodes.keys():
                        for stream in self.multiplex[node,'Material'][link].values():
                            process_carbon += stream['carbon_flow_rate']
                    if link[0] == 'PROD':
                        for stream in self.multiplex[node,'Material'][link].values():
                            process_carbon += stream['carbon_flow_rate']
                water = self.water_process(node)
                water_cons = water[0]-water[1]-water[2]-water[3]
                water_use[node] = water[0]

                if water_cons >= 0.1:
                    water_consumption[node] = water_cons
                try:
                    if process_carbon >= 0.1:
                        water_intensity[node] = water_cons/process_carbon
                except ZeroDivisionError:
                    pass

        # Remove processes that are not in both the steam consumption and steam intensity lists
        diff = set(water_intensity) - set(water_consumption)
        for x in diff:
            water_intensity.pop(x)

        diff = set(water_consumption) - set(water_intensity)
        for x in diff:
            water_consumption.pop(x)

        for process in ignore_list:
            water_consumption.pop(process)
            water_intensity.pop(process)
        # x and y given as array_like objects
        text = [x for x in water_consumption]

        fig = px.scatter(x=water_consumption,y=water_intensity,text=text, log_y=False, log_x=True, \
                        labels={'x':'Water consumption<br>kton CO2/year', \
                        'y':'Water intensity <br>[kton of water/(kton of carbon X year) ]'},\
                        width=width, height=height)
        fig.update_traces(textposition='top center', marker={'size':10,'color':'green'})
        fig.update_layout(font_family="Time New Roman", font_size=font_size)
        fig.show()


    def scatter_exergy(self, ignore_list="", height=1000, width =1000, font_size=10):
        """Method for plotting the steam intensity over the steam consumption in a scatter plot"""
        steam_consumption = {}
        steam_intensity = {}

        for node, node_value in self.process_nodes.items():
            process_energy = 0
            for name, value in node_value['steam_usage'].items():
                if value <= 0:
                    continue
                if name == 'LLPS':
                    process_energy += value * 1.970357667
                if name == 'LPS':
                    process_energy += value * 1.611524627
                if name == 'MPS':
                    process_energy += value * 1.495622128
                if name == 'HPS':
                    process_energy += value * 1.298335114
                if name == "HHPS":
                    process_energy += value * 1.24563498

            process_carbon = 0

            for link in self.multiplex[node,'Material']:
                if link[0] == node:
                    continue
                if "U" in link[0]:
                    # Ignore the carbon in streams send to utility units that
                    # is used as energy source
                    continue
                if link[0] in self.process_nodes.keys():
                    for stream in self.multiplex[node,'Material'][link].values():
                        process_carbon += stream['carbon_flow_rate']
                if link[0] == 'PROD':
                    for stream in self.multiplex[node,'Material'][link].values():
                        process_carbon += stream['carbon_flow_rate']

            if process_energy >= 0.1:
                steam_consumption[node] = process_energy

            try:
                if process_carbon >= 0.1:
                    steam_intensity[node] = process_energy/process_carbon
            except ZeroDivisionError:
                pass

        # Remove processes that are not in both the steam consumption and steam intensity lists
        diff = set(steam_intensity) - set(steam_consumption)
        for x in diff:
            steam_intensity.pop(x)
        diff = set(steam_consumption) - set(steam_intensity)
        for x in diff:
            steam_consumption.pop(x)

        for process in ignore_list:
            steam_consumption.pop(process)
            steam_intensity.pop(process)
        # x and y given as array_like objects
        text = [x for x in steam_consumption]

        fig = px.scatter(x=steam_consumption,y=steam_intensity,text=text, log_y=True, log_x=True,\
                        labels={'x':'Exergy<br>TJ/year',\
                        'y':'Exergy intensity <br>[TJ/(ktonne of carbon X year) ]'},\
                        width=width, height=height)
        fig.update_traces(textposition='top center', marker={'size':10})
        fig.update_layout(font_family="Time New Roman", font_size=font_size)
        fig.show()
        fig.write_image("scatter_exergy.pdf")
