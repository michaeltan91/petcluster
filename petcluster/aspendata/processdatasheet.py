'''Contains the methods for printing the mass and energy balances into an excel File in
the style agreed upon within the VICI project'''
from openpyxl import load_workbook
from openpyxl.styles import Font

class ProcessDataSheet(object):
    '''Main processdatasheet class for reporting Aspen Plus data to excel'''
    def __init__(self):

        self.temp = 1


    def print_mass(self, material_streams, excel_file):
        '''Report the material balance of the process in the format required for the
        process datasheet documentation as part of the VICI project'''

        sheet_name = 'Mass balances'
        work_book = load_workbook(excel_file)

        # if the sheet doesn't exist, create a new sheet
        try:
            sheet = work_book[sheet_name]
            work_book.remove(sheet)
            sheet = work_book.create_sheet(sheet_name)
        except KeyError:
            sheet = work_book.create_sheet(sheet_name)

        # Initialize counters
        count = 0
        feed_count, feed_mass = 0, 0
        product_count,product_mass = 0, 0
        waste_count, waste_mass = 0, 0

        start_count = 19

        # Iterate over all the material streams of the aspen model
        for name, obj in material_streams.items():
            comp_count = 0

            # If the material stream is a feed stream
            if obj.type == 'Feed':
                row_start = start_count + count
                # Iterate over all the components in the material stream and print the
                # component with a mass fraction larger than 0.0001
                for comp, frac in obj.massfrac.items():
                    if frac >= 0.0001:
                        sheet.cell(row=row_start + comp_count , column=4).value = comp
                        sheet.cell(row=row_start + comp_count , column=5).value = frac * 100
                        comp_count += 1
                        count += 1
                        feed_count += 1

                # Print the stream name, mass flowrate, temperature and pressure
                # Print stream name
                sheet.merge_cells(start_row= row_start, start_column=2,
                end_row=row_start+comp_count-1, end_column=2)
                sheet.cell(row= row_start , column=2).value = name
                # Print stream massflow
                sheet.merge_cells(start_row= row_start, start_column=3,
                end_row=row_start+comp_count-1, end_column=3)
                massflow = obj.massflow
                feed_mass += massflow
                sheet.cell(row= row_start , column=3).value = massflow
                # Print stream temperature
                sheet.merge_cells(start_row= row_start, start_column=6,
                end_row=row_start+comp_count-1, end_column=6)
                sheet.cell(row= row_start , column=6).value = obj.temperature
                # PRint stream pressure
                sheet.merge_cells(start_row= row_start, start_column=7,
                end_row=row_start+comp_count-1, end_column=7)
                sheet.cell(row= row_start , column=7).value = obj.pressure

            # If the material stream is a product stream
            if obj.type == 'Product':
                row_start = start_count + count + 3
                # Iterate over all the components in the material stream and print the
                # component with a mass fraction larger than 0.0001
                for comp, frac in obj.massfrac.items():
                    if frac >= 0.0001:

                        sheet.cell(row=row_start + comp_count , column=4).value = comp
                        sheet.cell(row=row_start + comp_count , column=5).value = frac * 100
                        comp_count += 1
                        count += 1
                        product_count += 1

                # Print the stream name, mass flowrate, temperature and pressure
                # Print stream name
                sheet.merge_cells(start_row= row_start, start_column=2,
                end_row=row_start+comp_count-1, end_column=2)
                sheet.cell(row= row_start , column=2).value = name
                # Print stream massflow
                sheet.merge_cells(start_row= row_start, start_column=3,
                end_row=row_start+comp_count-1, end_column=3)
                massflow = obj.massflow
                product_mass += massflow
                sheet.cell(row= row_start , column=3).value = massflow
                # Print stream temperature
                sheet.merge_cells(start_row= row_start, start_column=6,
                end_row=row_start+comp_count-1, end_column=6)
                sheet.cell(row= row_start , column=6).value = obj.temperature
                # Print stream pressure
                sheet.merge_cells(start_row= row_start, start_column=7,
                end_row=row_start+comp_count-1, end_column=7)
                sheet.cell(row= row_start , column=7).value = obj.pressure

            # If the material stream is a waste stream
            if obj.type == 'Waste':
                row_start = start_count + count + 6
                # Iterate over all the components in the material stream and print
                # the component with a mass fraction larger than 0.0001
                for comp, frac in obj.massfrac.items():
                    if frac >= 0.0001:
                        sheet.cell(row=row_start + comp_count , column=4).value = comp
                        sheet.cell(row=row_start + comp_count , column=5).value = frac * 100
                        comp_count += 1
                        waste_count += 1
                        count += 1

                # Print the stream name, mass flowrate, temperature and pressure
                # Print stream name
                sheet.merge_cells(start_row= row_start, start_column=2,
                end_row=row_start+comp_count-1, end_column=2)
                sheet.cell(row= row_start , column=2).value = name
                # Print stream mass flow
                sheet.merge_cells(start_row= row_start, start_column=3,
                end_row=row_start+comp_count-1, end_column=3)
                massflow = obj.massflow
                waste_mass += massflow
                sheet.cell(row= row_start , column=3).value = massflow
                # Print stream temperature
                sheet.merge_cells(start_row= row_start, start_column=6,
                end_row=row_start+comp_count-1, end_column=6)
                sheet.cell(row= row_start , column=6).value = obj.temperature
                # Print stream pressure
                sheet.merge_cells(start_row= row_start, start_column=7,
                end_row=row_start+comp_count-1, end_column=7)
                sheet.cell(row= row_start , column=7).value = obj.pressure

        # Print feed headers
        headers = ['Stream','Flowrate ktonne/y','Component','Concentration (wt%)',
        'Temperature ( C )','Pressure (bar)']
        feed_start = start_count
        col_count = 2
        self.fill_cell_bold(sheet, feed_start-3, col_count, 'Mass balances')

        self.fill_cell_bold(sheet, feed_start-2, col_count, 'Raw materials')
        for head in headers:
            self.fill_cell_bold(sheet, feed_start-1, col_count, head)
            col_count += 1

        # Print product headers
        prod_start = feed_start + feed_count + 3
        col_count = 2
        self.fill_cell_bold(sheet, prod_start-2, col_count, 'Products')
        for head in headers:
            self.fill_cell_bold(sheet, prod_start-1, col_count, head)
            col_count += 1

        # Print waste headers
        wast_start = prod_start + product_count + 3
        col_count = 2
        self.fill_cell_bold(sheet, wast_start-2, col_count, 'Waste streams')
        for head in headers:
            self.fill_cell_bold(sheet, wast_start-1, col_count, head)
            col_count += 1

        # Print Balance check
        check_start = wast_start + waste_count
        self.fill_cell_bold(sheet,check_start-7, 9,'Balance check')
        self.fill_cell_bold(sheet, check_start-6, 9,'Total inputs')
        self.fill_cell_bold(sheet, check_start-6, 10, 'Total F.rate ktonne/y ')
        self.fill_cell_bold(sheet, check_start-5, 9, 'Raw materials')
        sheet.cell(row=check_start-5, column= 10).value = feed_mass
        self.fill_cell_bold(sheet, check_start-4, 9, 'Products')
        sheet.cell(row=check_start-4, column= 10).value = product_mass
        self.fill_cell_bold(sheet, check_start-3, 9, 'Waste streams')
        sheet.cell(row=check_start-3, column= 10).value = waste_mass
        self.fill_cell_bold(sheet, check_start-2, 9, 'Balance check')
        sheet.cell(row=check_start-2, column= 10).value = feed_mass - product_mass - waste_mass
        self.fill_cell_bold(sheet, check_start-1, 9, 'Balance error')
        sheet.cell(row=check_start-1, column= 10).value = \
        (feed_mass - product_mass - waste_mass) / feed_mass

        work_book.save(excel_file)



    def print_energy(self, natural_gas, coolwater, electricity,
        refrigerant, steam, steam_gen, excel_file):
        '''Report the energy balance of the process in the format required for the
        process datasheet documentation as part of the VICI project'''

        sheet_name = 'Energy balances'
        work_book = load_workbook(excel_file)
        # if the sheet doesn't exist, create a new sheet
        try:
            sheet = work_book[sheet_name]
            work_book.remove(sheet)
            sheet = work_book.create_sheet(sheet_name)
        except KeyError:
            sheet = work_book.create_sheet(sheet_name)

        # Initialize counters and set headers for the left and right columns
        start = 18
        headleft = ['Utility type','Input - TJ/y', 'Output TJ/y', 'Net TJ/y', 'Remark']
        headright = ['Utility type', 'Inlet temperature (C)', 'Outlet temperature (C)',
        'Inlet pressure (bar)', 'Outlet pressure', 'Heat capacity/latent heat/lower heating',
        'Unit','Remark']

        count_col = 2
        self.fill_cell_bold(sheet, start-2, count_col, 'Overall energy needs')
        for head in headleft:
            self.fill_cell_bold(sheet, start-1, count_col, head)
            count_col += 1

        count_col = 10
        self.fill_cell_bold(sheet, start-2, count_col, 'Utility process conditions')
        for head in headright:
            self.fill_cell_bold(sheet, start-1, count_col, head)
            count_col += 1

        # Initialize the headers and counters for printing the utility per block
        util_head = ['Unit name','Unit description','Duty MJ/hr',
        'Duty TJ/y', 'Mass ktonne/y', 'Remark']
        elec_head = ['Unit name','Unit description','Power kW',
        'Energy GWh/y', 'Energy TJ/y', 'Remark']
        cntrl = 0
        cntrr = 0
        row_start = 18

        # Initialize list of the possible steam types of the VICI project
        steam_types = [['LLPS','LLPS-GEN'],['LPS','LPS-GEN'],['MPS','MPS-GEN'],
        ['HPS','HPS-GEN'],['HHPS','HHPS-GEN']]

        # Initialize list of the possible refrigerant types of the VICI project
        refrigerant_types = ['CHILLED','R134a', 'R717', 'R-410a', 'R41', 'R1150', 'R740']

        # Steam
        # Iterate over the list of steam levels
        for stm_type in steam_types:
            head_count = 0
            self.fill_cell_bold(sheet, row_start + cntrl, 2, stm_type[0])
            cntrl += 1
            for header in util_head:
                self.fill_cell_bold(sheet,  row_start + cntrl, 2 + head_count, header)
                head_count += 1
            cntrl += 1
            try:
                # For the selected steam level print the steam consumption duty and usage per block
                for block in steam[stm_type[0]].blocks:
                    sheet.cell(row= row_start + cntrl , column=2).value = block.uid
                    sheet.cell(row= row_start + cntrl, column=4).value = block.duty
                    sheet.cell(row= row_start + cntrl, column=5).value = block.duty * 8000 * 1E-6
                    sheet.cell(row= row_start + cntrl, column=6).value = block.usage
                    cntrl += 1
            except KeyError:
                cntrl += 1
            try:
                # For the selected steam level print the steam generation duty and usage per block
                for block in steam_gen[stm_type[1]].blocks:
                    sheet.cell(row= row_start + cntrl , column=2).value = block.uid
                    sheet.cell(row= row_start + cntrl, column=4).value = -block.duty
                    sheet.cell(row= row_start + cntrl, column=5).value = -block.duty * 8000 * 1E-6
                    sheet.cell(row= row_start + cntrl, column=6).value = -block.usage
                    cntrl += 1
            except KeyError:
                cntrl += 1
            cntrl += 1

        # Refrigerants
        for refrig_type in refrigerant_types:
            head_count = 0
            self.fill_cell_bold(sheet, row_start + cntrl, 2, refrig_type)
            cntrl += 1
            for header in util_head:
                self.fill_cell_bold(sheet,  row_start + cntrl, 2 + head_count, header)
                head_count += 1
            cntrl += 1
            try:
                # For the selected steam level print the steam consumption duty and usage per block
                for block in refrigerant[refrig_type].blocks:
                    sheet.cell(row= row_start + cntrl , column=2).value = block.uid
                    sheet.cell(row= row_start + cntrl, column=4).value = block.duty
                    sheet.cell(row= row_start + cntrl, column=5).value = block.duty * 8000 * 1E-6
                    sheet.cell(row= row_start + cntrl, column=6).value = block.usage
                    cntrl += 1
            except KeyError:
                pass
            cntrl += 1

        # Electricity
        head_count = 0
        self.fill_cell_bold(sheet, row_start + cntrr, 10, 'Electricity')
        cntrr += 1
        for header in elec_head:
            self.fill_cell_bold(sheet,  row_start + cntrr, 10 + head_count, header)
            head_count += 1
        cntrr += 1
        for util in electricity:
            # Print the electricity duty and usage per block
            for block in util.blocks:
                sheet.cell(row= row_start + cntrr , column=10).value = block.uid
                sheet.cell(row= row_start + cntrr, column=12).value = block.usage
                sheet.cell(row= row_start + cntrr, column=13).value = block.usage * 8000 * 1E-6
                sheet.cell(row= row_start + cntrr, column=14).value = \
                block.usage * 8000 * 1E-6 * 3.6
                cntrr += 1
        cntrr += 1

        # Natural gas
        head_count = 0
        self.fill_cell_bold(sheet, row_start + cntrr, 10, 'Natural gas')
        cntrr += 1
        for header in util_head:
            self.fill_cell_bold(sheet,  row_start + cntrr, 10 + head_count, header)
            head_count += 1
        cntrr += 1
        for util in natural_gas:
            # Print the natural gas duty and usage per block
            for block in util.blocks:
                sheet.cell(row= row_start + cntrr , column=10).value = block.uid
                sheet.cell(row= row_start + cntrr, column=12).value = block.duty
                sheet.cell(row= row_start + cntrr, column=13).value = block.duty * 8000 * 1E-6
                sheet.cell(row= row_start + cntrr, column=14).value = block.usage
                cntrr += 1
        cntrr += 1

        # Cooling water
        head_count = 0
        self.fill_cell_bold(sheet, row_start + cntrr, 10, 'Cooling water')
        cntrr += 1
        for header in util_head:
            self.fill_cell_bold(sheet,  row_start + cntrr, 10 + head_count, header)
            head_count += 1
        cntrr += 1
        for util in coolwater:
            # Print the cooling water duty and usage per block
            for block in util.blocks:
                sheet.cell(row= row_start + cntrr , column=10).value = block.uid
                sheet.cell(row= row_start + cntrr, column=12).value = block.duty
                sheet.cell(row= row_start + cntrr, column=13).value = block.duty * 1E-6 * 8000
                sheet.cell(row= row_start + cntrr, column=14).value = block.usage
                cntrr += 1
        cntrr += 1

        work_book.save(excel_file)


    def fill_cell(self, sheet, row1, column1, value):
        '''Fill an excel cell with a value'''
        data = sheet.cell(row=row1, column=column1)
        data.value = value

    def fill_cell_bold(self, sheet, row1, column1, value):
        '''Fill an excel with a value printed in bold'''
        data = sheet.cell(row=row1, column=column1)
        data.value = value
        data.font = Font(bold=True)
