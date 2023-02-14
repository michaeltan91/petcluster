'''Contains the data structure and retrieval from the Aspen Plus simulations'''
from warnings import warn
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from aspenauto import Model

from .stream import Stream
from .processdatasheet import ProcessDataSheet
from .utility import ElectricityManual, ManualUtility, NaturalGasManual, \
                     SteamGenManual, SteamStripping, HighTemperatureHeat

class Process(object):
    '''Main process class containing the data structure'''
    def __init__(self, aspen_file):
        self.energy = {}
        self.energy_use = {}
        self.energy_production = {}
        self.steam_usage = {}
        self.material_feed = {}
        self.material_product = {}
        self.material_waste = {}
        self.material_all = {}
        self.superstructure = {}

        # Create auxiliary dicts
        self.material_auxiliary = {}
        self.material_auxiliary['NG'] = {}
        self.material_auxiliary['H2'] = {}
        self.material_auxiliary['CO2'] = {}
        self.material_auxiliary['O2'] = {}
        self.material_auxiliary['AIR'] = {}

        # Load the Aspen model and run it
        self.aspen = Model(aspen_file)
        self.aspen.run(report_error=False)


    def add_manual_steam_gen(self, steam_type, block, heatstream, steam_stream_id):
        '''Add a manual steam generation utility'''
        # Retrieve the heatstream and material stream from the aspen simulation
        heat_stream = self.aspen.heat_streams[heatstream]
        mass_stream =self.aspen.streams[steam_stream_id]

        # Assign the steam generation utility to the requested block,
        steam = SteamGenManual(block, heat_stream, mass_stream )
        try:
            self.aspen.utilities[steam_type].blocks[block] = steam
            self.aspen.steam_gen[steam_type].blocks[block] = steam

        # When the requested utility is not yet defined, it is created
        except KeyError:
            utility = ManualUtility()
            self.aspen.utilities[steam_type] = utility
            self.aspen.steam_gen[steam_type] = utility
            self.aspen.utilities[steam_type].blocks[block] = steam
            self.aspen.steam_gen[steam_type].blocks[block] = steam
            warn(f"Created utility {steam_type} if not intended check syntax")


    def add_manual_steam_stripping(self, steam_type, block, stream_id):
        '''Add the thermal energy of steam stripping manually to the energy balance'''
        steam_stream = self.aspen.streams[stream_id]
        steam_stripping = SteamStripping(block, steam_stream, steam_type)

        try:
            self.aspen.utilities[steam_type].blocks[block] = steam_stripping
            self.aspen.steam[steam_type].blocks[block] = steam_stripping

        # If the requested utility is not yet defined, it is created
        except KeyError:
            utility = ManualUtility()
            self.aspen.utilities[steam_type] = utility
            self.aspen.steam[steam_type] = utility
            self.aspen.utilities[steam_type].blocks[block] = steam_stripping
            self.aspen.steam[steam_type].blocks[block] = steam_stripping
            warn(f"Created utility {steam_type} if not intended check syntax")


    def add_manual_natural_gas(self, block, ng_stream_id):
        '''Add a manual natural gas utility'''

        # Retrieve the natural gas material stream from the aspen simulation
        ng_stream = self.aspen.streams[ng_stream_id]

        natural_gas = NaturalGasManual(block, ng_stream)
        try:
            self.aspen.utilities['NG'].blocks[block] = natural_gas
            self.aspen.natural_gas['NG'].blocks[block] = natural_gas

        # If the requested utility is not yet defined, it is created
        except KeyError:
            utility = ManualUtility()
            self.aspen.utilities['NG'] = utility
            self.aspen.natural_gas['NG'] = utility
            self.aspen.utilities['NG'].blocks[block] = natural_gas
            self.aspen.natural_gas['NG'].blocks[block] = natural_gas
            warn("Created utility NG if not intended check syntax")


    def add_manual_electricity(self, block, elec_stream_id):
        '''Add a manual electricity utility'''

        elec_stream = self.aspen.streams[elec_stream_id]
        electricity = ElectricityManual(block, elec_stream)
        try:
            self.aspen.utilities['ELECTRIC'].blocks[block] = electricity
            self.aspen.electricity['ELECTRIC'].blocks[block] = electricity
        except KeyError:
            utility = ManualUtility()
            self.aspen.utilities['ELECTRIC'] = utility
            self.aspen.electricity['ELECTRIC'] = utility
            self.aspen.utilities['ELECTRIC'].blocks[block] = electricity
            self.aspen.electricity['ELECTRIC'].blocks[block] = electricity
            warn('Created utility ELECTRIC if not intended check syntax')


    def add_high_temperature_heat(self, block, heatstream_id):
        """Add a high temperature heat_stream"""

        heatstream = self.aspen.streams[heatstream_id]

        heat = HighTemperatureHeat(block, heatstream)
        try:
            self.aspen.utilities['HTHEAT'].blocks[block] = heat
            self.aspen.htheat['HTHEAT'].blocks[block] = heat
        except KeyError:
            utility = ManualUtility()
            self.aspen.utilities['HTHEAT'] = utility
            self.aspen.htheat['HTHEAT'] = utility
            self.aspen.utilities['HTHEAT'].blocks[block] = heat
            self.aspen.htheat['HTHEAT'].blocks[block] = heat


    def load_process_data(self):
        '''
        Load the process data from the Aspen Model
        Add the manual utilities before loading the complete process data!!!!
        '''

        outlet_streams = []
        feed_streams = []
        product_streams = []
        waste_streams = []

        #self.utilities = self.aspen.utilities

        # Fill lists for material feed, product and waste streams
        for stream in self.aspen.material_streams:
            if stream.type == "Feed":
                feed_streams.append(stream.name)
            elif stream.type == "Product":
                product_streams.append(stream.name)
                outlet_streams.append(stream.name)
            elif stream.type == "Waste":
                waste_streams.append(stream.name)
                outlet_streams.append(stream.name)

        # Load material feed data from Aspen model
        for stream_id in feed_streams:
            stream = Stream(self.aspen.streams[stream_id])
            self.material_feed[stream_id] = stream
            self.material_all[stream_id] = stream
            if self.aspen.streams[stream_id].auxiliary == 'NG':
                self.material_auxiliary['NG'][stream_id] = stream
            elif self.aspen.streams[stream_id].auxiliary == 'H2':
                self.material_auxiliary['H2'][stream_id] = stream
            elif self.aspen.streams[stream_id].auxiliary == 'CO2':
                self.material_auxiliary['CO2'][stream_id] = stream
            elif self.aspen.streams[stream_id].auxiliary == 'O2':
                self.material_auxiliary['O2'][stream_id] = stream
            elif self.aspen.streams[stream_id].auxiliary == 'AIR':
                self.material_auxiliary['AIR'][stream_id] = stream
        # Load material product data from Aspen model
        for stream_id in product_streams:
            stream = Stream(self.aspen.streams[stream_id])
            self.material_product[stream_id] = stream
            self.material_all[stream_id] = stream
        # Load material waste data from Aspen model
        for stream_id in waste_streams:
            stream = Stream(self.aspen.streams[stream_id])
            self.material_waste[stream_id] = stream
            self.material_all[stream_id] = stream

        # Make a list of the standard steam definitions of the VICI project
        steam_types = [['LLPS','LLPS-GEN'], ['LPS','LPS-GEN'],['MPS','MPS-GEN'],
        ['HPS','HPS-GEN'], ['HHPS','HHPS-GEN']]

        # Load the total steam duty per type for the process
        for steam in steam_types:
            temp = 0
            temp2 = 0

            duty_use = 0
            duty_prod = 0
            mass_use = 0
            mass_prod = 0

            # Load the total steam duty being utilized by the process
            try:
                for block in self.aspen.utilities[steam[0]].blocks:
                    temp += block.duty
                    temp2 += block.usage

                    duty_use += block.duty
                    mass_use += block.usage
            except KeyError:
                pass
            # Load the total steam duty being generated by the process
            try:
                for block in self.aspen.utilities[steam[1]].blocks:
                    temp -= block.duty
                    temp2 -= block.usage

                    duty_prod += block.duty
                    mass_prod += block.usage
            except KeyError:
                pass
            # Set the units of energy to TJ per operating year
            self.energy[steam[0]] = temp * 8000 * 1E-6
            self.steam_usage[steam[0]] = temp2

            self.energy_use[steam[0]] = [duty_use* 8000 * 1E-6, mass_use]
            self.energy_production[steam[1]] = [duty_prod* 8000 * 1E-6, mass_prod]


        # Load the total electricity duty being utilized by the process
        temp = 0
        try:
            for block in self.aspen.electricity['ELECTRIC'].blocks:
                temp += block.usage
            self.energy['Electricity'] = temp* 8000 * 1E-6 * 3.6
        except KeyError:
            pass

        # Load the total cooling water duty being utilized by the process
        temp = 0
        try:
            for block in self.aspen.coolwater['CW'].blocks:
                temp += block.duty
            self.energy['Cooling water'] = temp * 8000 * 1E-6
        except KeyError:
            pass

        # Load the total natural gas duty being utilized by the process
        temp = 0
        try:
            for block in self.aspen.natural_gas['NG'].blocks:
                temp += block.duty
            self.energy['NG'] = temp * 8000 * 1E-6
        except KeyError:
            pass

        # Load chilled water
        temp = 0
        try:
            for block in self.aspen.refrigerant['CHILLED'].blocks:
                temp += block.duty
            self.energy['Chilled water'] = temp * 8000 * 1E-6
        except KeyError:
            pass

        # Load refrigerants
        refrigerant_types = ['R134A', 'R717', 'R-410A', 'R41', 'R1150', 'R740','LIN']
        for refrig in refrigerant_types:
            temp = 0
            try:
                for block in self.aspen.refrigerant[refrig].blocks:
                    temp += block.duty
                self.energy[refrig] = temp * 8000 * 1E-6
            except KeyError:
                pass


        header = pd.MultiIndex.from_product([outlet_streams,['m_out','massflow','X']],
        names=['stream', 'type'])
        data_frame = pd.DataFrame(np.zeros([0,len(header)]),
                        columns=header)

        for name in outlet_streams:
            count = 0
            for comp, value in self.aspen.streams[name].massfrac.items():
                if value > 0.001:
                    data_frame.loc[count,(name,'m_out')] = comp
                    data_frame.loc[count,(name,'massflow')] = \
                    value * self.aspen.streams[name].massflow
                    count += 1
        self.superstructure = data_frame


    def report(self, excel_file):
        '''Reports the mass and energy balance of the process into an excel  sheet'''
        pds = ProcessDataSheet()
        pds.print_mass(self.aspen.streams, excel_file)
        pds.print_energy(self.aspen.natural_gas, self.aspen.coolwater, self.aspen.electricity,
        self.aspen.refrigerant, self.aspen.steam, self.aspen.steam_gen, excel_file)


    def close_aspen(self):
        """Closes the Aspen model and shuts down the Engine"""
        self.aspen.close()
        del self.aspen


    def e_factor(self):
        """Calculate the environmental factor a process, defined as the mass of waste
        (excluding water) divided by the total mass of product"""
        product_mass = 0
        waste_mass = 0

        # Determine the total product mass
        for _, stream in self.material_feed.items():
            product_mass += stream.massflow

        # Determine the total waste mass
        for _, stream in self.material_waste.items():
            waste_mass += stream.massflow
            # Exclude the contribution of water if part of the waste stream
            try:
                waste_mass -= stream.massfrac['H2O'] * stream.massflow
            except KeyError:
                pass

        return waste_mass / product_mass


    def calculate_carbon_fraction(self, component_dict):
        """Calculate the carbon fraction of all the material feed, product and waste streams"""
        if isinstance(component_dict, pd.DataFrame) is False:
            component_dict = pd.read_excel(component_dict, index_col=1)

        feed, product, waste = 0, 0, 0
        for _, stream in self.material_feed.items():
            stream.calc_carbon_frac(component_dict)
            feed += stream.massflow * stream.carbonfrac

        for _, stream in self.material_product.items():
            stream.calc_carbon_frac(component_dict)
            product += stream.massflow * stream.carbonfrac

        for _, stream in self.material_waste.items():
            stream.calc_carbon_frac(component_dict)
            waste += stream.massflow * stream.carbonfrac

        try:
            diff = abs(feed-product-waste)/feed
            if diff >= 0.0001:
                print('carbon leak')
        except ZeroDivisionError:
            pass



    def carbon_intensity(self, component_dict):
        '''Returns the carbon intensity of the process, the ratio of kg of carbon in the product
        and the kg of carbon in the feed'''
        carbon_feed, carbon_waste, carbon_product = 0, 0, 0
        for _, stream in self.material_feed.items():
            temp = 0

            for component, value in stream.molefrac.items():
                temp += stream.moleflow * value * component_dict['Carbon Atoms'][component] * \
                12.01 *8000*1E-6
            stream.carbonfrac = temp/stream.massflow
            carbon_feed += temp

        for _, stream in self.material_product.items():
            temp = 0
            for component, value in stream.molefrac.items():
                temp += stream.moleflow * value * component_dict['Carbon Atoms'][component] * \
                12.01 *8000*1E-6
            stream.carbonfrac = temp/stream.massflow
            carbon_product += temp

        for _, stream in self.material_waste.items():
            temp = 0
            for component, value in stream.molefrac.items():
                temp += stream.moleflow * value * component_dict['Carbon Atoms'][component] * \
                12.01 *8000*1E-6
            stream.carbonfrac = temp/stream.massflow
            carbon_waste += temp

        balance = carbon_feed - carbon_product - carbon_waste
        if balance/carbon_feed > 0.0001:
            warn("Carbon leak in process")
        else:
            return min(carbon_product/carbon_feed, 1)


    def process_water(self):
        '''Returns the total amount of water used in the process excluding cooling water'''
        water = 0
        # Determine the total amount of water used in the process
        for _, stream in self.material_feed.items():
            try:
                water += stream.massfrac['H2O'] * stream.massflow
            except KeyError:
                pass

        return water


    def collect_energy(self, excel_file):
        """Collects all energy data  from the Aspen Plus model and saves in an Excel file"""
        energy_duty = {}
        energy_usage = {}

        sheet_name = "name"
        work_book = load_workbook(excel_file)

        # if the sheet doesn't exist, create a new sheet
        try:
            sheet = work_book[sheet_name]
            work_book.remove(sheet)
            sheet = work_book.create_sheet(sheet_name)
        except KeyError:
            sheet = work_book.create_sheet(sheet_name)

        # Make a list of the standard steam definitions of the VICI project
        steam_types = [['LLPS','LLPS-GEN'], ['LPS','LPS-GEN'],['MPS','MPS-GEN'],
        ['HPS','HPS-GEN'], ['HHPS','HHPS-GEN']]

        # Load the total steam duty per type for the process
        for steam in steam_types:
            temp1 = 0
            temp2 = 0
            # Load the total steam duty being utilized by the process
            try:
                for block in self.aspen.utilities[steam[0]].blocks:
                    temp1 += block.duty
                    temp2 += block.usage
            except KeyError:
                pass
            # Load the total steam duty being generated by the process
            try:
                for block in self.aspen.utilities[steam[1]].blocks:
                    temp1 -= block.duty
                    temp2 -= block.usage
            except KeyError:
                pass
            # Set the units of energy to TJ per operating year
            energy_duty[steam[0]] = temp1 / 3600
            energy_usage[steam[0]] = temp2

        i = 0
        for steam in steam_types:
            sheet.cell(row= i+1, column=1).value = steam[0]
            sheet.cell(row= i+1, column=2).value = energy_duty[steam[0]]
            sheet.cell(row= i+1, column=3).value = energy_usage[steam[0]]

            i += 1

        work_book.save(excel_file)
