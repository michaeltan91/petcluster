from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill

class ProcessDataSheet(object):

    def __init__(self):

        self.temp = 1


    def Print_Mass(self, material_streams, work_book):
        
        sheet_name = 'Mass balances' 
        
        wb = load_workbook(work_book)
    
        # if the sheet doesn't exist, create a new sheet
        try:
            sheet = wb[sheet_name]
            wb.remove(sheet)
            sheet = wb.create_sheet(sheet_name)
        except:
            sheet = wb.create_sheet(sheet_name)

        count = 0
        feed_count, feed_mass = 0, 0
        product_count,product_mass = 0, 0
        waste_count, waste_mass = 0, 0

        
        rw = 19

        for name, obj in material_streams.items():
            comp_count = 0

            if obj.type == 'Feed':
                row_start = rw + count
                for comp, frac in obj.massfrac.items():
                
                    if frac >= 0.0001:
                        
                        sheet.cell(row=row_start + comp_count , column=4).value = comp
                        sheet.cell(row=row_start + comp_count , column=5).value = frac * 100
                        comp_count += 1
                        count += 1
                        feed_count += 1

                sheet.merge_cells(start_row= row_start, start_column=2, end_row=row_start+comp_count-1, end_column=2)
                sheet.cell(row= row_start , column=2).value = name
                sheet.merge_cells(start_row= row_start, start_column=3, end_row=row_start+comp_count-1, end_column=3)
                massflow = obj.massflow
                feed_mass += massflow
                sheet.cell(row= row_start , column=3).value = massflow
                sheet.merge_cells(start_row= row_start, start_column=6, end_row=row_start+comp_count-1, end_column=6)
                sheet.cell(row= row_start , column=6).value = obj.temperature
                sheet.merge_cells(start_row= row_start, start_column=7, end_row=row_start+comp_count-1, end_column=7)
                sheet.cell(row= row_start , column=7).value = obj.pressure


            if obj.type == 'Product':
                row_start = rw + count + 3
                for comp, frac in obj.massfrac.items():
                    if frac >= 0.0001:

                        sheet.cell(row=row_start + comp_count , column=4).value = comp
                        sheet.cell(row=row_start + comp_count , column=5).value = frac * 100
                        comp_count += 1
                        count += 1
                        product_count += 1

                sheet.merge_cells(start_row= row_start, start_column=2, end_row=row_start+comp_count-1, end_column=2)
                sheet.cell(row= row_start , column=2).value = name
                sheet.merge_cells(start_row= row_start, start_column=3, end_row=row_start+comp_count-1, end_column=3)
                massflow = obj.massflow
                product_mass += massflow
                sheet.cell(row= row_start , column=3).value = massflow
                sheet.merge_cells(start_row= row_start, start_column=6, end_row=row_start+comp_count-1, end_column=6)
                sheet.cell(row= row_start , column=6).value = obj.temperature
                sheet.merge_cells(start_row= row_start, start_column=7, end_row=row_start+comp_count-1, end_column=7)
                sheet.cell(row= row_start , column=7).value = obj.pressure


            if obj.type == 'Waste':
                row_start = rw + count + 6
                for comp, frac in obj.massfrac.items():
                    if frac >= 0.0001:
                        sheet.cell(row=row_start + comp_count , column=4).value = comp
                        sheet.cell(row=row_start + comp_count , column=5).value = frac * 100
                        comp_count += 1
                        waste_count += 1
                        count += 1

                sheet.merge_cells(start_row= row_start, start_column=2, end_row=row_start+comp_count-1, end_column=2)
                sheet.cell(row= row_start , column=2).value = name
                sheet.merge_cells(start_row= row_start, start_column=3, end_row=row_start+comp_count-1, end_column=3)
                massflow = obj.massflow
                waste_mass += massflow
                sheet.cell(row= row_start , column=3).value = massflow
                sheet.merge_cells(start_row= row_start, start_column=6, end_row=row_start+comp_count-1, end_column=6)
                sheet.cell(row= row_start , column=6).value = obj.temperature
                sheet.merge_cells(start_row= row_start, start_column=7, end_row=row_start+comp_count-1, end_column=7)
                sheet.cell(row= row_start , column=7).value = obj.pressure
        
        # Print feed headers
        headers = ['Stream','Flowrate ktonne/y','Component','Concentration (wt%)','Temperature ( C )','Pressure (bar)']
        feed_start = rw
        col_count = 2
        self.fill_cell_bold(sheet, feed_start-3, col_count, 'Mass balances')

        self.fill_cell_bold(sheet, feed_start-2, col_count, 'Raw materials')
        for head in headers:
            self.fill_cell_bold(sheet, feed_start-1, col_count, head)
            col_count += 1

        # Print product headers 
        prod_start = feed_start + feed_count + 3
        col_count = 2
        self.fill_cell_bold(sheet, prod_start-2, col_count, 'Products')
        for head in headers:
            self.fill_cell_bold(sheet, prod_start-1, col_count, head)
            col_count += 1

        # Print waste headers
        wast_start = prod_start + product_count + 3
        col_count = 2
        self.fill_cell_bold(sheet, wast_start-2, col_count, 'Waste streams')
        for head in headers:
            self.fill_cell_bold(sheet, wast_start-1, col_count, head)
            col_count += 1
        
        # Print Balance check
        check_start = wast_start + waste_count
        self.fill_cell_bold(sheet,check_start-7, 9,'Balance check')
        self.fill_cell_bold(sheet, check_start-6, 9,'Total inputs')
        self.fill_cell_bold(sheet, check_start-6, 10, 'Total F.rate ktonne/y ')
        self.fill_cell_bold(sheet, check_start-5, 9, 'Raw materials')
        sheet.cell(row=check_start-5, column= 10).value = feed_mass
        self.fill_cell_bold(sheet, check_start-4, 9, 'Products')
        sheet.cell(row=check_start-4, column= 10).value = product_mass
        self.fill_cell_bold(sheet, check_start-3, 9, 'Waste streams')
        sheet.cell(row=check_start-3, column= 10).value = waste_mass
        self.fill_cell_bold(sheet, check_start-2, 9, 'Balance check')
        sheet.cell(row=check_start-2, column= 10).value = feed_mass - product_mass - waste_mass
        self.fill_cell_bold(sheet, check_start-1, 9, 'Balance error %')
        sheet.cell(row=check_start-1, column= 10).value = (feed_mass - product_mass - waste_mass)/feed_mass

        wb.save(work_book)



    def Print_Energy(self, natural_gas, coolwater, electricity,
        refrigerant, steam, steam_gen, work_book):
        
        sheet_name = 'Energy balances' 
        
        wb = load_workbook(work_book)
        try:
            sheet = wb[sheet_name]
            wb.remove(sheet)
            sheet = wb.create_sheet(sheet_name)
        except:
            sheet = wb.create_sheet(sheet_name)

        start = 18
        util_count = 0
        headleft = ['Utility type','Input - TJ/y', 'Output TJ/y', 'Net TJ/y', 'Remark']
        headright = ['Utility type', 'Inlet temperature (C)', 'Outlet temperature (C)', 'Inlet pressure (bar)', 
        'Outlet pressure', 'Heat capacity/latent heat/lower heating', 'Unit','Remark']
        Utility = ['Low low pressure steam', 'Low pressure steam', 'Medium pressure steam', 'High pressure steam', 
        'Refrigerant1', 'Refrigerant2', 'Refrigerant3', 'Refrigerant4', 'Electricity', 'Natural gas','Cooling water']
        
        count_col = 2
        self.fill_cell_bold(sheet, start-2, count_col, 'Overall energy needs')
        for head in headleft:
            self.fill_cell_bold(sheet, start-1, count_col, head)
            count_col += 1
        
        count_col = 10
        self.fill_cell_bold(sheet, start-2, count_col, 'Utility process conditions')
        for head in headright:
            self.fill_cell_bold(sheet, start-1, count_col, head)
            count_col += 1
        
        util_head = ['Unit name','Unit description','Duty MJ/hr', 'Duty TJ/y', 'Mass ktonne/y', 'Remark']
        elec_head = ['Unit name','Unit description','Power kW', 'Energy GWh/y', 'Energy TJ/y', 'Remark']
        cntrl = 0 
        cntrr = 0
        row_start = 18


        steam_types = [['LLPS','LLPS-GEN'],['LPS','LPS-GEN'],['MPS','MPS-GEN'],['HPS','HPS-GEN'],['HHPS','HHPS-GEN']]

        
        
        for stm_type in steam_types:
            head_count = 0
            self.fill_cell_bold(sheet, row_start + cntrl, 2, stm_type[0])
            cntrl += 1
            for header in util_head:
                self.fill_cell_bold(sheet,  row_start + cntrl, 2 + head_count, header)
                head_count += 1
            cntrl += 1
            try:
                for block in steam[stm_type[0]].blocks:
                    sheet.cell(row= row_start + cntrl , column=2).value = block.uid
                    sheet.cell(row= row_start + cntrl, column=4).value = block.duty
                    sheet.cell(row= row_start + cntrl, column=5).value = block.duty * 8000 * 1E-6
                    sheet.cell(row= row_start + cntrl, column=6).value = block.usage
                    cntrl += 1
            except KeyError:
                cntrl += 1
                pass
            try:
                for block in steam_gen[stm_type[1]].blocks:
                    sheet.cell(row= row_start + cntrl , column=2).value = block.uid
                    sheet.cell(row= row_start + cntrl, column=4).value = -block.duty
                    sheet.cell(row= row_start + cntrl, column=5).value = -block.duty * 8000 * 1E-6
                    sheet.cell(row= row_start + cntrl, column=6).value = -block.usage
                    cntrl += 1
            except KeyError:
                cntrl += 1
                pass
            cntrl += 1


      
        head_count = 0
        self.fill_cell_bold(sheet, row_start + cntrr, 10, 'Electricity')
        cntrr += 1
        for header in elec_head:
            self.fill_cell_bold(sheet,  row_start + cntrr, 10 + head_count, header)
            head_count += 1
        cntrr += 1
        for util in electricity:
            for block in util.blocks:
                sheet.cell(row= row_start + cntrr , column=10).value = block.uid
                sheet.cell(row= row_start + cntrr, column=12).value = block.usage
                sheet.cell(row= row_start + cntrr, column=13).value = block.usage * 8000 * 1E-6
                sheet.cell(row= row_start + cntrr, column=14).value = block.usage * 8000 * 1E-6 * 3.6
                cntrr += 1
        cntrr += 1
        
        head_count = 0
        self.fill_cell_bold(sheet, row_start + cntrr, 10, 'Natural gas')
        cntrr += 1
        for header in util_head:
            self.fill_cell_bold(sheet,  row_start + cntrr, 10 + head_count, header)
            head_count += 1
        cntrr += 1
        for util in natural_gas:
            for block in util.blocks:
                sheet.cell(row= row_start + cntrr , column=10).value = block.uid
                sheet.cell(row= row_start + cntrr, column=12).value = block.duty
                sheet.cell(row= row_start + cntrr, column=13).value = block.duty * 8000 * 1E-6
                sheet.cell(row= row_start + cntrr, column=14).value = block.usage
                cntrr += 1
        cntrr += 1

       
        head_count = 0
        self.fill_cell_bold(sheet, row_start + cntrr, 10, 'Cooling water')
        cntrr += 1
        for header in util_head:
            self.fill_cell_bold(sheet,  row_start + cntrr, 10 + head_count, header)
            head_count += 1
        cntrr += 1
        for util in coolwater:
            for block in util.blocks:
                sheet.cell(row= row_start + cntrr , column=10).value = block.uid
                sheet.cell(row= row_start + cntrr, column=12).value = block.duty
                sheet.cell(row= row_start + cntrr, column=13).value = block.duty * 1E-6 * 8000
                sheet.cell(row= row_start + cntrr, column=14).value = block.usage
                cntrr += 1
        cntrr += 1
    

        wb.save(work_book)

    
    def fill_cell(self, sheet, row1, column1, value):
        sheet.cell(row=row1, column=column1)
    
    def fill_cell_bold(self, sheet, row1, column1, value):
        c = sheet.cell(row=row1, column=column1)
        c.value = value
        c.font = Font(bold=True)